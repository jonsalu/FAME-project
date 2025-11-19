from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from io import BytesIO
import zipfile
import io

# 🧹 Funções auxiliares seguras
def normalize_text(v):
    """Remove espaços e padroniza texto, mesmo se for None."""
    if v is None:
        return ""
    return str(v).strip().upper()


async def handleDoubleData(file, columns_to_check=None):
    print("➡️ Columns to check recebidas:", columns_to_check)

    # Carrega a planilha
    contents = await file.read()
    workbook = load_workbook(BytesIO(contents))
    sheet = workbook.active

    print("🔎 Colunas encontradas na planilha:", 
          [sheet.cell(row=1, column=c).value for c in range(1, sheet.max_column + 1)])

    # Se não vier nenhuma coluna, usa a primeira
    if not columns_to_check:
        columns_to_check = [sheet.cell(row=1, column=1).value]

    # 🎨 Cor de destaque
    amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # 🔎 Mapeia nomes de cabeçalhos → índices de coluna
    header_map = {}
    for col in range(1, sheet.max_column + 1):
        header_value = normalize_text(sheet.cell(row=1, column=col).value)
        header_map[header_value] = col

    print("📋 Header map:", header_map)

    # Converte os nomes de colunas recebidos em índices (1-based)
    col_indexes = []
    for col_name in columns_to_check:
        norm = normalize_text(col_name)
        if norm in header_map:
            col_indexes.append(header_map[norm])

    print("📊 Colunas encontradas:", col_indexes)

    if not col_indexes:
        workbook.close()
        raise ValueError("Nenhum dos cabeçalhos selecionados foi encontrado na planilha.")

    # -------------------------------------------------------
    # 🔥 NOVO: detectar duplicatas pela COMBINAÇÃO das colunas
    # -------------------------------------------------------

    combinations = {}         # combo_key → [row_numbers]
    duplicate_rows = set()    # linhas duplicadas pela combinação

    for i in range(2, sheet.max_row + 1):
        # cria a chave da combinação (tuple)
        combo_key = tuple(
            normalize_text(sheet.cell(row=i, column=col).value)
            for col in col_indexes
        )

        # ignora combinações totalmente vazias
        if all(v == "" for v in combo_key):
            continue

        if combo_key not in combinations:
            combinations[combo_key] = [i]
        else:
            combinations[combo_key].append(i)

    # pega todas as linhas onde teve duplicata (combinação aparece > 1)
    for rows in combinations.values():
        if len(rows) > 1:
            duplicate_rows.update(rows)

    # -------------------------------------------------------
    # 🔶 MARCAR LINHA INTEIRA (todas as colunas)
    # -------------------------------------------------------
    for row in duplicate_rows:
        for col in range(1, sheet.max_column + 1):
            sheet.cell(row=row, column=col).fill = amarelo

    # 💾 Salva o resultado
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()

    # 📦 Cria um ZIP com o arquivo tratado
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zipf:
        zipf.writestr("duplicados_tratados.xlsx", output.getvalue())

    zip_buffer.seek(0)
    return zip_buffer.getvalue()
