import io
import zipfile
import datetime
import unicodedata
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# --- Helpers de Normalização ---

def remover_acentos(texto):
    """Remove acentos e caracteres especiais de uma string."""
    if not texto:
        return ""
    # NFD separa o caractere da acentuação (ex: 'á' -> 'a' + '´')
    # Mn filtra apenas as marcas de acentuação
    return "".join(
        c for c in unicodedata.normalize('NFD', texto)
        if unicodedata.category(c) != 'Mn'
    )

def normalize(value):
    """Normaliza o valor para comparação: remove acentos, trim e uppercase."""
    if value is None: 
        return ""
    
    # Se for número inteiro em formato float (ex: 10.0), trata como 10
    if isinstance(value, (int, float)) and value == int(value): 
        value = int(value)
    
    texto = str(value).strip().upper()
    return remover_acentos(texto)

def sanitize(value):
    """Prepara o valor para exibição no JSON (ISO para datas, string para o resto)."""
    if isinstance(value, (datetime.date, datetime.datetime)): 
        return value.isoformat()
    return value if value is not None else ""

# --- FUNÇÃO ÚNICA DE COMPARAÇÃO ---
async def comparar_planilhas(file1, file2, modo="download"):
    # 1. Carregar arquivos em memória
    b1 = await file1.read()
    b2 = await file2.read()
    
    wb1 = load_workbook(io.BytesIO(b1))
    wb2 = load_workbook(io.BytesIO(b2))
    sheet1 = wb1.active
    sheet2 = wb2.active

    # Estilos de cores
    verde = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    vermelho = PatternFill(start_color="F08080", end_color="F08080", fill_type="solid")

    # 2. Configurar Colunas (Dinâmico)
    max_col = max(sheet1.max_column, sheet2.max_column)
    cols = list(range(1, max_col + 1))

    # 3. Indexar Planilha 2 para busca rápida (usando a chave sem acentos)
    index2 = {}
    for r in range(2, sheet2.max_row + 1):
        key = tuple(normalize(sheet2.cell(row=r, column=c).value) for c in cols)
        index2.setdefault(key, []).append(r)

    matched_rows_sheet2 = set()
    preview_data = [] 

    # 4. PROCESSAR PLANILHA 1 (Comparação com a P2)
    for r in range(2, sheet1.max_row + 1):
        key = tuple(normalize(sheet1.cell(row=r, column=c).value) for c in cols)
        
        match_rows = index2.get(key, [])
        status = "vermelho"
        
        if match_rows:
            status = "verde"
            # Remove o índice para evitar que duplicatas na P1 usem a mesma linha da P2
            r2 = match_rows.pop(0)
            matched_rows_sheet2.add(r2)

        # Ação: Pintar (Somente se for modo download para economizar CPU)
        if modo == "download":
            fill = verde if status == "verde" else vermelho
            for c in cols:
                sheet1.cell(row=r, column=c).fill = fill

        # Ação: Gerar JSON (Somente se for modo preview)
        if modo == "preview":
            row_json = {"origem": "PLANILHA_1", "campos": {}}
            for c in cols:
                header = str(sheet1.cell(row=1, column=c).value or f"C{c}")
                val = sanitize(sheet1.cell(row=r, column=c).value)
                row_json["campos"][header] = {"valor": val, "status": status}
            preview_data.append(row_json)

    # 5. PROCESSAR PLANILHA 2 (Verificar o que sobrou)
    for r in range(2, sheet2.max_row + 1):
        status = "verde" if r in matched_rows_sheet2 else "vermelho"

        if modo == "download":
            fill = verde if status == "verde" else vermelho
            for c in cols:
                sheet2.cell(row=r, column=c).fill = fill
            
        if modo == "preview":
            row_json = {"origem": "PLANILHA_2", "campos": {}}
            for c in cols:
                header = str(sheet2.cell(row=1, column=c).value or f"C{c}")
                val = sanitize(sheet2.cell(row=r, column=c).value)
                row_json["campos"][header] = {"valor": val, "status": status}
            preview_data.append(row_json)

    # 6. RETORNO FINAL
    if modo == "preview":
        return {"dados": preview_data}

    # Se for modo download, gera os arquivos e zips
    out1 = io.BytesIO()
    out2 = io.BytesIO()
    wb1.save(out1)
    wb2.save(out2)
    
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("planilha_base_verificada.xlsx", out1.getvalue())
        z.writestr("planilha_nova_verificada.xlsx", out2.getvalue())
    
    zip_buffer.seek(0)
    return zip_buffer