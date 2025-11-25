from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from io import BytesIO
import zipfile
import io
import datetime

# Reutilizando sua função auxiliar (mantenha ela no escopo)
def normalize_text(v):
    if v is None:
        return ""
    return str(v).strip().upper()

async def processAndMarkMissing(file, columns_to_check=None, modo="download"):
    print(f"➡️ Buscando vazios. Modo: {modo}")

    # 1. Carrega a planilha (Exatamente igual ao seu código anterior)
    contents = await file.read()
    workbook = load_workbook(BytesIO(contents))
    sheet = workbook.active

    # Se não vier nenhuma coluna, usa todas ou a primeira (regra de negócio sua)
    # Aqui vou assumir que se não vier, checa todas
    if not columns_to_check:
        columns_to_check = [sheet.cell(row=1, column=c).value for c in range(1, sheet.max_column + 1)]

    # 2. Cor de destaque (Vermelho para erro)
    vermelho = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

    # 3. Mapeia cabeçalhos (Igual ao seu código)
    header_map = {}
    for col in range(1, sheet.max_column + 1):
        header_value = normalize_text(sheet.cell(row=1, column=col).value)
        header_map[header_value] = col

    # Converte nomes para índices
    col_indexes = []
    for col_name in columns_to_check:
        norm = normalize_text(col_name)
        if norm in header_map:
            col_indexes.append(header_map[norm])

    if not col_indexes:
        workbook.close()
        raise ValueError("Nenhum dos cabeçalhos selecionados foi encontrado.")

    # -------------------------------------------------------
    # 🔥 LÓGICA DE DETECÇÃO (A única parte que muda de verdade)
    # -------------------------------------------------------
    missing_rows = set()

    for i in range(2, sheet.max_row + 1):
        # Verifica se ALGUMA das colunas selecionadas está vazia nesta linha
        is_row_missing = False
        for col_idx in col_indexes:
            val = sheet.cell(row=i, column=col_idx).value
            # Considera vazio se for None ou string vazia/espaços
            if val is None or str(val).strip() == "":
                is_row_missing = True
                break # Se achou um vazio na linha, já marca e pula para a próxima lógica
        
        if is_row_missing:
            missing_rows.add(i)

    # -------------------------------------------------------
    # 👁️ MODO PREVIEW (Cópia exata da sua lógica de duplicados)
    # -------------------------------------------------------
    if modo == "preview":
        preview_list = []

        # Ordena para mostrar na ordem que aparecem no Excel
        for row in sorted(missing_rows):
            row_data = {}

            for col in range(1, sheet.max_column + 1):
                header = sheet.cell(row=1, column=col).value
                value = sheet.cell(row=row, column=col).value

                # 🔥 Convertendo tudo para tipos JSON-safe (Sua lógica segura)
                if isinstance(value, datetime.datetime):
                    value = value.isoformat()
                elif isinstance(value, (datetime.date, datetime.time)):
                    value = str(value)
                elif value is None:
                    value = None
                else:
                    value = value if isinstance(value, (int, float, bool, str)) else str(value)

                row_data[header] = value

            preview_list.append(row_data)

        workbook.close()
        return preview_list # Retorna Lista de Dicts (JSON)

    # -------------------------------------------------------
    # ⬇️ MODO DOWNLOAD (Pinta e baixa)
    # -------------------------------------------------------
    for row in missing_rows:
        for col in range(1, sheet.max_column + 1):
            sheet.cell(row=row, column=col).fill = vermelho

    output = io.BytesIO()
    workbook.save(output)
    workbook.close()

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zipf:
        zipf.writestr("linhas_vazias_marcadas.xlsx", output.getvalue())

    zip_buffer.seek(0)
    return zip_buffer.getvalue() # Retorna Bytes